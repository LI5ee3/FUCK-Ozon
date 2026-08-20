import csv
import io
from datetime import datetime, timezone

from openpyxl import load_workbook

from .db import transaction

CHANNELS = {"FBP", "realFBS", "WHD"}
SHIPPED_STATUS_PARTS = ("已签收", "运输中", "待取件", "已送达", "已完成", "已发货")


def _text(value):
    return "" if value is None else str(value).strip()


def _number(value):
    text = _text(value).replace(" ", "").replace(",", ".")
    return float(text) if text else None


def _utc(value):
    text = _text(value)
    if not text:
        return None
    dt = datetime.fromisoformat(text.replace("Z", "+00:00"))
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc).isoformat().replace("+00:00", "Z")


def _shipping(row):
    status = _text(row.get("状态"))
    shipped_at = _text(row.get("已转移配送") or row.get("实际转移配送日期"))
    status_shipped = any(part in status for part in SHIPPED_STATUS_PARTS)
    shipped = bool(shipped_at or status_shipped)
    cancelled = status == "已取消"
    anomaly = bool(shipped_at) != status_shipped
    return int(shipped), (int(shipped) if cancelled else None), int(anomaly), shipped_at


def import_csv(shop_id, channel, filename, content):
    if channel not in CHANNELS:
        raise ValueError("未知渠道")
    rows = list(csv.DictReader(io.StringIO(content.decode("utf-8-sig")), delimiter=";"))
    required = {"订单号", "发货号码", "状态", "SKU", "数量"}
    if not rows or not required.issubset(rows[0]):
        raise ValueError("CSV字段不符合Ozon导出格式")

    quantities = {}
    for row in rows:
        posting = _text(row["发货号码"])
        sku = _text(row["SKU"])
        if not posting or not sku:
            raise ValueError("订单号或SKU为空")
        quantity = _number(row["数量"])
        if quantity is None or quantity <= 0 or not quantity.is_integer():
            raise ValueError(f"订单 {posting} 的数量无效")
        key = posting, sku
        quantities[key] = quantities.get(key, 0) + int(quantity)

    with transaction() as db:
        batch = db.execute(
            "INSERT INTO import_batches(shop_id,kind,filename,row_count) VALUES(?,?,?,?)",
            (shop_id, channel, filename, len(rows)),
        ).lastrowid
        for row in rows:
            posting = _text(row["发货号码"])
            sku = _text(row["SKU"])
            quantity = quantities[(posting, sku)]
            shipped, cancelled_after, anomaly, shipped_at = _shipping(row)
            created = row.get("已创建") or row.get("正在处理中")
            amount = _number(row.get("发货的金额"))
            amount_currency = _text(row.get("货件的货币代码")) or None
            db.execute("""
              INSERT INTO orders(shop_id,posting_number,parent_order_no,channel,created_at,shipped_at,
                delivered_at,status_raw,cancel_reason_raw,shipped,cancelled_after_ship,data_anomaly,
                amount_original,amount_currency,buyer_paid,buyer_currency,source,import_batch_id)
              VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
              ON CONFLICT(shop_id,posting_number) DO UPDATE SET
                parent_order_no=COALESCE(orders.parent_order_no,excluded.parent_order_no),
                channel=CASE WHEN orders.source IN ('api','push') THEN orders.channel ELSE excluded.channel END,
                created_at=COALESCE(orders.created_at,excluded.created_at),
                shipped_at=COALESCE(orders.shipped_at,excluded.shipped_at),
                delivered_at=COALESCE(orders.delivered_at,excluded.delivered_at),
                status_raw=CASE WHEN orders.source IN ('api','push') AND orders.status_raw<>'' THEN orders.status_raw ELSE excluded.status_raw END,
                cancel_reason_raw=COALESCE(orders.cancel_reason_raw,excluded.cancel_reason_raw),
                shipped=CASE WHEN orders.source IN ('api','push') AND NOT (orders.channel='WHD' AND orders.status_raw='已取消')
                  THEN orders.shipped ELSE excluded.shipped END,
                cancelled_after_ship=COALESCE(orders.cancelled_after_ship,excluded.cancelled_after_ship),
                data_anomaly=MAX(orders.data_anomaly,excluded.data_anomaly),
                amount_original=COALESCE(orders.amount_original,excluded.amount_original),
                amount_currency=COALESCE(orders.amount_currency,excluded.amount_currency),
                buyer_paid=COALESCE(orders.buyer_paid,excluded.buyer_paid),
                buyer_currency=COALESCE(orders.buyer_currency,excluded.buyer_currency),
                updated_at=strftime('%Y-%m-%dT%H:%M:%SZ','now')
            """, (shop_id, posting, _text(row["订单号"]), channel, _utc(created), _utc(shipped_at),
                  _utc(row.get("实际配送日期") or row.get("配送日期")), _text(row["状态"]),
                  _text(row.get("取消原因")) or None, shipped, cancelled_after, anomaly,
                  amount, amount_currency, _number(row.get("已由买家支付")),
                  _text(row.get("买家货币代码")) or None, "csv", batch))
            item_channel = db.execute(
                "SELECT channel FROM orders WHERE shop_id=? AND posting_number=?", (shop_id, posting)).fetchone()[0]
            db.execute("UPDATE order_items SET channel=? WHERE shop_id=? AND posting_number=?",
                       (item_channel, shop_id, posting))
            db.execute("""
              INSERT INTO order_items(shop_id,channel,posting_number,sku,offer_id,product_name_raw,
                quantity,unit_price,price_currency,buyer_paid,buyer_currency,source,import_batch_id)
              VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)
              ON CONFLICT(shop_id,posting_number,sku) DO UPDATE SET
                channel=excluded.channel,offer_id=COALESCE(order_items.offer_id,excluded.offer_id),
                product_name_raw=CASE WHEN order_items.source IN ('api','push') AND order_items.product_name_raw<>'' THEN order_items.product_name_raw ELSE excluded.product_name_raw END,
                quantity=CASE WHEN order_items.source IN ('api','push') THEN order_items.quantity ELSE excluded.quantity END,
                unit_price=COALESCE(order_items.unit_price,excluded.unit_price),
                price_currency=COALESCE(order_items.price_currency,excluded.price_currency),
                buyer_paid=COALESCE(order_items.buyer_paid,excluded.buyer_paid),
                buyer_currency=COALESCE(order_items.buyer_currency,excluded.buyer_currency),
                import_batch_id=CASE WHEN order_items.source IN ('api','push') THEN order_items.import_batch_id ELSE excluded.import_batch_id END
            """, (shop_id, item_channel, posting, sku, _text(row.get("货号")) or None,
                  _text(row.get("商品名称")), quantity,
                  _number(row.get("您的价格")), _text(row.get("商品的货币代码")) or None,
                  _number(row.get("已由买家支付")), _text(row.get("买家货币代码")) or None,
                  "csv", batch))
    return {"batch_id": batch, "rows": len(rows)}


def import_costs(shop_id, filename, content):
    book = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = book.active
    rows = sheet.iter_rows(values_only=True)
    headers = [_text(v) for v in next(rows, ())]
    required = {"订单编号", "汇率(原币)", "商品总成本"}
    if not required.issubset(headers):
        raise ValueError("XLSX字段不符合马帮导出格式")
    indexes = {name: headers.index(name) for name in required}
    values = [row for row in rows if _text(row[indexes["订单编号"]])]
    with transaction() as db:
        batch = db.execute(
            "INSERT INTO import_batches(shop_id,kind,filename,row_count) VALUES(?,?,?,?)",
            (shop_id, "马帮成本", filename, len(values)),
        ).lastrowid
        for row in values:
            cost = _number(row[indexes["商品总成本"]])
            if cost is None or cost < 0:
                raise ValueError(f"订单 {_text(row[indexes['订单编号']])} 的成本无效")
            db.execute("""
              INSERT INTO order_costs(shop_id,posting_number,cost_cny,source_rate,source,import_batch_id)
              VALUES(?,?,?,?,?,?) ON CONFLICT(shop_id,posting_number) DO UPDATE SET
                cost_cny=excluded.cost_cny,source_rate=excluded.source_rate,import_batch_id=excluded.import_batch_id
            """, (shop_id, _text(row[indexes["订单编号"]]),
                  cost, _number(row[indexes["汇率(原币)"]]),
                  "mabang", batch))
    return {"batch_id": batch, "rows": len(values)}
