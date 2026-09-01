import json
import math
import re
from datetime import date, datetime, time, timezone
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from urllib.parse import urlparse
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from .db import connect, transaction


REQUIRED_HEADERS = (
    "订单编号", "平台SKU", "平台SKU数量", "平台SKU单个成本", "汇率(原币)", "平台链接",
)
COMPARE_FIELDS = ("offer_id", "quantity", "unit_cost", "exchange_rate_original", "platform_link")
MAX_SQLITE_INTEGER = 2**63 - 1


def _text(value):
    return "" if value is None else str(value).strip()


def _raw_value(value):
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, float) and not math.isfinite(value):
        return str(value)
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _raw_payload(headers, values):
    return json.dumps(
        {header: _raw_value(values[index]) for index, header in enumerate(headers)
         if header and index < len(values)},
        ensure_ascii=False, separators=(",", ":"), allow_nan=False,
    )


def _decimal(value, label, *, required=False, positive=False):
    if value is None or (type(value) is str and not value.strip()):
        if required:
            raise ValueError(f"{label}不能为空")
        return None, None
    if type(value) is bool:
        raise ValueError(f"{label}必须为数字")
    try:
        number = Decimal(str(value).strip())
    except (InvalidOperation, TypeError, ValueError, OverflowError) as error:
        raise ValueError(f"{label}必须为有限数字") from error
    if not number.is_finite() or number < 0 or (positive and number <= 0):
        qualifier = "有限正数字" if positive else "有限非负数字"
        raise ValueError(f"{label}必须为{qualifier}")
    return number, format(number, "f")


def _quantity(value):
    if value is None or (type(value) is str and not value.strip()) or type(value) is bool:
        raise ValueError("平台SKU数量必须为正整数")
    try:
        number = Decimal(str(value).strip())
    except (InvalidOperation, TypeError, ValueError, OverflowError) as error:
        raise ValueError("平台SKU数量必须为正整数") from error
    if (not number.is_finite() or number <= 0 or number != number.to_integral_value()
            or number > MAX_SQLITE_INTEGER):
        raise ValueError("平台SKU数量必须为正整数")
    return int(number)


def _ozon_sku(platform_link):
    try:
        parsed = urlparse(platform_link.strip())
        hostname = (parsed.hostname or "").lower()
    except ValueError:
        return None
    if (parsed.scheme not in ("http", "https")
            or (hostname != "ozon.ru" and not hostname.endswith(".ozon.ru"))):
        return None
    segment = parsed.path.rstrip("/").rsplit("/", 1)[-1]
    match = re.search(r"(\d+)$", segment)
    return match.group(1) if match else None


def _find_header(worksheet):
    for row_no, values in enumerate(
            worksheet.iter_rows(min_row=1, max_row=20, values_only=True), 1):
        headers = [_text(value) for value in values]
        named_headers = [header for header in headers if header]
        if len(named_headers) != len(set(named_headers)):
            continue
        if set(REQUIRED_HEADERS).issubset(named_headers):
            return row_no, headers, {header: headers.index(header) for header in REQUIRED_HEADERS}
    return None


def _has_value(values):
    return any(value is not None and (not isinstance(value, str) or value.strip()) for value in values)


def _parse_row(headers, positions, values, source_row_no):
    def value(header):
        index = positions[header]
        return values[index] if index < len(values) else None

    raw_payload_json = _raw_payload(headers, values)
    erp_order_number = _text(value("订单编号"))
    if not erp_order_number:
        raise ValueError("订单编号不能为空")

    offer_id = _text(value("平台SKU")) or None
    quantity = _quantity(value("平台SKU数量"))
    unit_cost_number, unit_cost = _decimal(value("平台SKU单个成本"), "平台SKU单个成本", required=True)
    _, exchange_rate_original = _decimal(
        value("汇率(原币)"), "汇率(原币)", positive=True,
    )
    platform_link = "" if value("平台链接") is None else str(value("平台链接"))
    ozon_sku = _ozon_sku(platform_link)
    if not ozon_sku:
        raise ValueError("平台链接无法解析 Ozon SKU")

    return {
        "erp_order_number": erp_order_number,
        "ozon_sku": ozon_sku,
        "offer_id": offer_id,
        "quantity": quantity,
        "unit_cost": unit_cost,
        "exchange_rate_original": exchange_rate_original,
        "total_cost": format(unit_cost_number * quantity, "f"),
        "platform_link": platform_link,
        "source_row_no": source_row_no,
        "raw_payload_json": raw_payload_json,
    }


def _parse_workbook(workbook):
    for worksheet in workbook.worksheets:
        header = _find_header(worksheet)
        if not header:
            continue
        header_row, headers, positions = header
        records = {}
        row_count = 0
        for source_row_no, values in enumerate(
                worksheet.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
            if not _has_value(values):
                continue
            row_count += 1
            try:
                record = _parse_row(headers, positions, values, source_row_no)
            except ValueError as error:
                raise ValueError(f"第{source_row_no}行：{error}") from error
            key = (record["erp_order_number"], record["ozon_sku"])
            previous = records.get(key)
            if previous is None:
                records[key] = record
                continue
            if all(previous[field] == record[field] for field in COMPARE_FIELDS[:-1]):
                continue
            rows = f"{previous['source_row_no']}、{source_row_no}"
            raise ValueError(
                f"ERP订单 {key[0]}、Ozon SKU {key[1]} 在第{rows}行存在冲突"
            )
        return row_count, list(records.values())
    raise ValueError("Excel字段不符合马帮 ERP 成本导出格式")


def _utc_now():
    return datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")


def import_erp_costs(shop_id, filename, content):
    if Path(filename).suffix.lower() != ".xlsx":
        raise ValueError("仅支持XLSX文件")
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except (BadZipFile, InvalidFileException, OSError, ValueError) as error:
        raise ValueError("Excel字段不符合马帮 ERP 成本导出格式") from error
    try:
        row_count, records = _parse_workbook(workbook)
    finally:
        workbook.close()

    imported_at = _utc_now()
    inserted_count = updated_count = unchanged_count = 0
    with transaction() as db:
        batch_id = db.execute("""
          INSERT INTO erp_cost_import_batches(
            shop_id,filename,row_count,parsed_count,inserted_count,updated_count,unchanged_count,imported_at)
          VALUES(?,?,?,?,?,?,?,?)
        """, (shop_id, filename, row_count, len(records), 0, 0, 0, imported_at)).lastrowid
        for record in records:
            existing = db.execute("""
              SELECT offer_id,quantity,unit_cost,exchange_rate_original,platform_link
              FROM erp_order_item_costs
              WHERE shop_id=? AND erp_order_number=? AND ozon_sku=?
            """, (shop_id, record["erp_order_number"], record["ozon_sku"])).fetchone()
            if existing is None:
                inserted_count += 1
            elif all(existing[field] == record[field] for field in COMPARE_FIELDS):
                unchanged_count += 1
            else:
                updated_count += 1
            db.execute("""
              INSERT INTO erp_order_item_costs(
                shop_id,erp_order_number,ozon_sku,offer_id,quantity,unit_cost,
                exchange_rate_original,total_cost,platform_link,source_batch_id,
                source_row_no,raw_payload_json,imported_at,updated_at)
              VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)
              ON CONFLICT(shop_id,erp_order_number,ozon_sku) DO UPDATE SET
                offer_id=excluded.offer_id, quantity=excluded.quantity, unit_cost=excluded.unit_cost,
                exchange_rate_original=excluded.exchange_rate_original, total_cost=excluded.total_cost,
                platform_link=excluded.platform_link, source_batch_id=excluded.source_batch_id,
                source_row_no=excluded.source_row_no, raw_payload_json=excluded.raw_payload_json,
                updated_at=excluded.updated_at
            """, (shop_id, record["erp_order_number"], record["ozon_sku"], record["offer_id"],
                  record["quantity"], record["unit_cost"], record["exchange_rate_original"],
                  record["total_cost"], record["platform_link"], batch_id, record["source_row_no"],
                  record["raw_payload_json"], imported_at, imported_at))
        db.execute("""
          UPDATE erp_cost_import_batches
          SET inserted_count=?,updated_count=?,unchanged_count=?
          WHERE id=?
        """, (inserted_count, updated_count, unchanged_count, batch_id))
    return {
        "batch_id": batch_id, "rows": row_count, "parsed": len(records),
        "inserted": inserted_count, "updated": updated_count, "unchanged": unchanged_count,
    }


def list_erp_cost_imports():
    with connect() as db:
        return [dict(row) for row in db.execute("""
          SELECT b.*,s.name shop_name
          FROM erp_cost_import_batches b JOIN shops s ON s.id=b.shop_id
          ORDER BY b.id DESC LIMIT 10
        """)]
